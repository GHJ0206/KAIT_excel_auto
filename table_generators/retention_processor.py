"""
표명 : 기록물철별 보존기간 책정 현황

기능 :
기록물분류기준표 Excel 파일을 분석하여
보존기간별 기록물철 개수와 비율을 자동 산출한다.
"""

import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment

NORMALIZE = {

    "1": "1년",
    "1년": "1년",

    "3": "3년",
    "3년": "3년",

    "5": "5년",
    "5년": "5년",

    "10": "10년",
    "10년": "10년",

    "30": "30년",
    "30년": "30년",

    "준영구": "준영구",

    "영구": "영구"
}

# input data
def build_input_path(institution, year):

    file_path = (
        f"input_data/기록물분류기준표_{institution}_{year}.xlsx"
    )

    return file_path


def get_sheet_names(file_path):

    workbook = load_workbook(file_path)

    sheet_names = []

    for sheet in workbook.worksheets:
        sheet_names.append(sheet.title)

    return workbook, sheet_names


def select_sheet(workbook):

    while True:

        try:

            choice = int(
                input(
                    "분석할 시트 번호를 입력해주세요: "
                )
            )

            # 범위 검사
            if 1 <= choice <= len(workbook.worksheets):

                selected_sheet = (
                    workbook.worksheets[choice - 1]
                )

                return selected_sheet

            else:
                print("잘못된 번호입니다.")

        except ValueError:

            print("숫자만 입력해주세요.")


def find_retention_column(sheet):

    retention_col = None
    header_row = None

    # 1행 ~ 10행 탐색
    for row in range(1, 11):

        # 모든 열 탐색
        for col in range(1, sheet.max_column + 1):

            value = sheet.cell(
                row=row,
                column=col
            ).value

            if value is None:
                continue

            value = str(value).strip()

            if value == "보존기간":

                retention_col = col
                header_row = row

                return retention_col, header_row

    return None, None


def classify_retention_values(
        sheet,
        retention_col,
        header_row
):

    counts = {

        "1년": 0,
        "3년": 0,
        "5년": 0,
        "10년": 0,
        "30년": 0,
        "준영구": 0,
        "영구": 0,

        "복합": 0,

        "미책정": 0
    }

    # 헤더 아래부터 마지막 행까지 탐색
    for row in range(
            header_row + 2,
            sheet.max_row + 1
    ):

        # 행 전체가 비어있는지 확인
        row_has_data = False

        for col in range(1, sheet.max_column + 1):

            cell_value = sheet.cell(
                row=row,
                column=col
            ).value

            if cell_value is not None:
                row_has_data = True
                break

        # 완전히 빈 행이면 무시
        if not row_has_data:
            continue

        value = sheet.cell(
            row=row,
            column=retention_col
        ).value

        # CASE 1 : 빈 셀
        if value is None:

            print("미책정 행 확인:", row)

            for col in range(1, sheet.max_column + 1):
                print(
                    "열",
                    col,
                    ":",
                    sheet.cell(row=row, column=col).value
                )

            counts["미책정"] += 1

            continue
            print("빈 셀 발견 행번호:", row)

            counts["미책정"] += 1

            continue

        value = str(value).strip()

        # CASE 2 : 빈 문자열
        if value == "":

            counts["미책정"] += 1

            continue

        # CASE 3 : 정상값 확인
        normalized = NORMALIZE.get(value)

        if normalized is not None:

            counts[normalized] += 1

        # CASE 4 : 그 외 = 복합
        else:

            counts["복합"] += 1

    return counts


def create_headers(counts):

    headers = [

        "구분",

        "1년",
        "3년",
        "5년",
        "10년",
        "30년",
        "준영구",
        "영구"
    ]

    # 복합 존재하면 추가
    if counts["복합"] > 0:
        headers.append("복합")

    # 미책정 존재하면 추가
    if counts["미책정"] > 0:
        headers.append("미책정")

    # 마지막은 항상 합계
    headers.append("합계")

    return headers


def create_output_excel(
        counts,
        headers,
        institution,
        year
):

    wb = Workbook()

    ws = wb.active

    # 헤더 작성
    for idx, header in enumerate(headers, start=1):

        ws.cell(
            row=1,
            column=idx,
            value=header
        )

    # 첫 번째 열
    ws["A2"] = "기록물철"
    ws["A3"] = "비율"

    # 합계 계산
    total = sum(counts.values())

    # 데이터 입력
    col_index = 2

    for header in headers[1:]:

        # 합계
        if header == "합계":

            ws.cell(
                row=2,
                column=col_index,
                value=total
            )

            ws.cell(
                row=3,
                column=col_index,
                value="100%"
            )

        else:

            count = counts[header]

            ws.cell(
                row=2,
                column=col_index,
                value=count
            )

            # 비율 계산
            if total == 0:

                ratio = 0

            else:

                ratio = round(
                    (count / total) * 100,
                    1
                )

            ws.cell(
                row=3,
                column=col_index,
                value=f"{ratio}%"
            )

        col_index += 1

    # 가운데 정렬
    center = Alignment(
        horizontal="center",
        vertical="center"
    )

    for row in range(1, 4):

        for col in range(
                1,
                len(headers) + 1
        ):

            ws.cell(
                row=row,
                column=col
            ).alignment = center

    # 천 단위 쉼표
    for col in range(
            2,
            len(headers) + 1
    ):

        ws.cell(
            row=2,
            column=col
        ).number_format = "#,##0"

    # 저장 경로
    output_file = (

        f"output_data/"
        f"기록물철별 보존기간 책정 현황"
        f"({year})_{institution}.xlsx"
    )

    wb.save(output_file)

    print("파일 생성 완료")

    print(output_file)


def process_excel(file_path, sheet_name):
    pass


def run_local():

    print("===== 기록물철별 보존기간 책정 현황 =====")

    # 기관명 입력
    institution = input(
        "분석할 기관명을 띄어쓰기 없이 정확하게 입력해주세요: "
    )

    # 기준년도 입력
    year = input(
        "기준년도를 한글 없이 숫자만 입력해주세요: "
    )

    # 파일 경로 생성
    file_path = build_input_path(
        institution,
        year
    )

    print("찾을 파일:", file_path)

    # 파일 존재 여부 확인
    if not os.path.exists(file_path):

        print("입력하신 파일이 존재하지 않습니다.")

        return

    print("파일을 찾았습니다.")

    # workbook 열기 + 시트 목록 읽기
    workbook, sheet_names = get_sheet_names(file_path)

    print("===== 시트 목록 =====")

    for idx, sheet_name in enumerate(sheet_names, start=1):

        print(f"{idx}. {sheet_name}")

        # 시트 선택
        selected_sheet = select_sheet(workbook)

        print("선택한 시트:", selected_sheet.title)

        # 보존기간 컬럼 찾기
        retention_col, header_row = (
            find_retention_column(selected_sheet)
        )

        # 못 찾은 경우
        if retention_col is None:
            print(
                "'보존기간' 컬럼을 찾을 수 없습니다."
            )

            return

        print("보존기간 컬럼을 찾았습니다.")

        print("헤더 행 번호:", header_row)

        print("헤더 열 번호:", retention_col)

        # 보존기간 분류
        counts = classify_retention_values(
            selected_sheet,
            retention_col,
            header_row
        )

        print("===== 집계 결과 =====")

        for key, value in counts.items():

            print(key, ":", value)

        headers = create_headers(counts)

        print("===== 생성될 헤더 =====")

        print(headers)

        create_output_excel(
            counts,
            headers,
            institution,
            year
        )